from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook

class Excel():
    def __init__(self, file_name, *args):
        self.file_name = file_name
        self.args = args
        
    def get_cell_value(self):
        wb = load_workbook(filename=self.file_name, read_only=True)
        ws = wb['OD матрица']
        value = str(ws[self.args[0]].value)
        return value

    def get_sum(self):
        wb = load_workbook(filename=self.file_name, read_only=True)
        ws = wb['OD матрица']
        sum = 0
        for i in range(len(self.args)):
            value = int(ws[self.args[i]].value)
            sum += value
        return str(sum)



class Get_Image(Excel):
    def __init__(self, file_name, coordinate_x, coordinate_y, *args):
        super().__init__(file_name, *args)
        self.coordinate_x = coordinate_x
        self.coordinate_y = coordinate_y

    def draw_image(self):
        image = Image.open('Шаблон.jpg')
        font = ImageFont.truetype("arial.ttf", 20)
        drawer = ImageDraw.Draw(image)
        if len(self.args) == 1:
            drawer.text((self.coordinate_x, self.coordinate_y), Excel.get_cell_value(self), font=font, fill='black')
        else:
            drawer.text((self.coordinate_x, self.coordinate_y), Excel.get_sum(self), font=font, fill='#404040')
        image.save('Шаблон.jpg')

        


value1 = Get_Image('Мира - Бассейная (4).xlsx', 830, 213, 'B6') # С ул.Бассейная север - просп.Мира запад
value1.draw_image()

value2 = Get_Image('Мира - Бассейная (4).xlsx', 920, 120, 'E6') # С ул.Бассейная север - ул.Бассейная юг 
value2.draw_image()

value3 = Get_Image('Мира - Бассейная (4).xlsx', 997, 213, 'C6') # С ул.Бассейная север - просп.Мира восток 
value3.draw_image()

value4 = Get_Image('Мира - Бассейная (4).xlsx', 1096, 300, 'D5') # С просп.Мира восток - ул.Бассейная север 
value4.draw_image()

value5 = Get_Image('Мира - Бассейная (4).xlsx', 1096, 359, 'B5') # С просп.Мира восток - просп.Мира запад 
value5.draw_image()

value6 = Get_Image('Мира - Бассейная (4).xlsx', 1096, 460, 'E5') # С просп.Мира восток - ул.Бассейная юг 
value6.draw_image()

value7 = Get_Image('Мира - Бассейная (4).xlsx', 997, 546, 'C7') # С ул.Бассейная юг - просп.Мира восток
value7.draw_image()

value8 = Get_Image('Мира - Бассейная (4).xlsx', 915, 640, 'D7') # С ул.Бассейная юг - ул.Бассейная север 
value8.draw_image()

value9 = Get_Image('Мира - Бассейная (4).xlsx', 835, 546, 'B7') # С ул.Бассейная юг - просп.Мира запад 
value9.draw_image()

value10 = Get_Image('Мира - Бассейная (4).xlsx', 700, 460, 'E4') # С просп.Мира запад - ул.Бассейная юг 
value10.draw_image()

value11 = Get_Image('Мира - Бассейная (4).xlsx', 700, 359, 'C4') # С просп.Мира запад - просп.Мира восток 
value11.draw_image()

value12 = Get_Image('Мира - Бассейная (4).xlsx', 700, 300, 'D4') # С просп.Мира запад - ул.Бассейная север  
value12.draw_image()

sum1 = Get_Image('Мира - Бассейная (4).xlsx', 815, 81, 'C6', 'B6', 'E6') # Общий поток с ул.Бассейная север
sum1.draw_image()

sum2 = Get_Image('Мира - Бассейная (4).xlsx', 1005, 81, 'D4', 'D5', 'D7') # Общий поток на ул.Бассейная север
sum2.draw_image() 

sum3 = Get_Image('Мира - Бассейная (4).xlsx', 1225, 270, 'B5', 'C5', 'D5', 'E5') # Общий поток с просп.Мира восток
sum3.draw_image() 

sum4 = Get_Image('Мира - Бассейная (4).xlsx', 1225, 483, 'C4','C5', 'C7', 'C6') # Общий поток на просп.Мира восток
sum4.draw_image() 

sum5 = Get_Image('Мира - Бассейная (4).xlsx', 1000, 640, 'B7', 'D7', 'C7') # Общий поток с ул.Бассейная юг
sum5.draw_image() 

sum6 = Get_Image('Мира - Бассейная (4).xlsx', 825, 640, 'E4', 'E5', 'E6') # Общий поток на ул.Бассейная юг
sum6.draw_image() 

sum7 = Get_Image('Мира - Бассейная (4).xlsx', 585, 483, 'B4','C4', 'D4', 'E4') # Общий поток с просп.Мира запад 
sum7.draw_image() 

sum8 = Get_Image('Мира - Бассейная (4).xlsx', 585, 270, 'B4','B5', 'B6','B7') # Общий поток на просп.Мира запад 
sum8.draw_image() 


    



